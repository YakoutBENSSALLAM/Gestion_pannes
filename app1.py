from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, make_response
import sqlite3
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing, Rect
import io
import os

app = Flask(__name__)
app.secret_key = 'votre_cle_secrete_2025'

# Configuration du logo et thème dessalement
LOGO_PATH = 'static/images/logo.png'  # Chemin vers votre logo
COMPANY_NAME = "ONEE-BO"
COMPANY_SUBTITLE = "Système de Gestion des pannes"

# Couleurs thème dessalement (bleu océan)
THEME_COLORS = {
    'primary': '1E3A8A',      # Bleu foncé
    'secondary': '3B82F6',     # Bleu moyen
    'accent': '0EA5E9',        # Bleu clair
    'light': 'E0F2FE',         # Bleu très clair
    'success': '10B981',       # Vert
    'warning': 'F59E0B',       # Orange
    'danger': 'EF4444',        # Rouge
    'header': '0F172A'         # Bleu très foncé
}

def get_db_connection():
    conn = sqlite3.connect('gmao.db')
    conn.row_factory = sqlite3.Row
    return conn

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Accès refusé. Droits administrateur requis.', 'error')
            return redirect(url_for('historique'))
        return f(*args, **kwargs)
    return decorated_function

def create_excel_styles(wb):
    """Créer des styles Excel professionnels pour ONEE-BO"""
    
    # Style pour l'en-tête principal (plus élégant)
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')  # Gris anthracite
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(
        left=Side(style='thin', color='374151'),
        right=Side(style='thin', color='374151'),
        top=Side(style='thin', color='374151'),
        bottom=Side(style='thin', color='374151')
    )
    
    # Style pour les sous-en-têtes (plus sophistiqué)
    subheader_style = NamedStyle(name="subheader_style")
    subheader_style.font = Font(name='Calibri', size=13, bold=True, color='1F2937')
    subheader_style.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    subheader_style.alignment = Alignment(horizontal='center', vertical='center')
    subheader_style.border = Border(
        left=Side(style='medium', color='6B7280'),
        right=Side(style='medium', color='6B7280'),
        top=Side(style='medium', color='6B7280'),
        bottom=Side(style='medium', color='6B7280')
    )
    
    # Style pour les données importantes
    important_style = NamedStyle(name="important_style")
    important_style.font = Font(name='Calibri', size=11, bold=True, color='1F2937')
    important_style.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    important_style.alignment = Alignment(horizontal='left', vertical='center')
    important_style.border = Border(
        left=Side(style='thin', color='DBEAFE'),
        right=Side(style='thin', color='DBEAFE'),
        top=Side(style='thin', color='DBEAFE'),
        bottom=Side(style='thin', color='DBEAFE')
    )
    
    # Style pour les labels
    label_style = NamedStyle(name="label_style")
    label_style.font = Font(name='Calibri', size=11, bold=True, color='374151')
    label_style.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    label_style.alignment = Alignment(horizontal='right', vertical='center')
    label_style.border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Ajouter les styles au workbook
    try:
        wb.add_named_style(header_style)
        wb.add_named_style(subheader_style)
        wb.add_named_style(important_style)
        wb.add_named_style(label_style)
    except ValueError:
        pass
    
    return wb

def add_excel_header(ws, title, subtitle=None):
    """Ajouter un en-tête professionnel avec logo à gauche"""
    
    # Ajuster la hauteur des lignes pour un meilleur espacement
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 8
    
    # Logo à gauche (si disponible)
    if os.path.exists(LOGO_PATH):
        try:
            logo = OpenpyxlImage(LOGO_PATH)
            logo.height = 80
            logo.width = 80
            # Placer le logo à gauche
            ws.add_image(logo, 'A1')
        except Exception as e:
            print(f"Erreur lors de l'ajout du logo: {e}")
    
    # Titre principal de l'entreprise (décalé pour éviter le logo)
    ws['C1'] = COMPANY_NAME
    ws['C1'].font = Font(name='Arial', size=22, bold=True, color='1F2937')
    ws['C1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('C1:J1')
    
    # Sous-titre avec style élégant (décalé)
    ws['C2'] = COMPANY_SUBTITLE
    ws['C2'].font = Font(name='Arial', size=11, italic=True, color='6B7280')
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('C2:J2')
    
    # Ligne de séparation élégante
    for col in range(1, 11):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    
    # Titre du rapport avec encadré
    ws['A4'] = title
    ws['A4'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].border = Border(
        left=Side(style='medium', color='374151'),
        right=Side(style='medium', color='374151'),
        top=Side(style='medium', color='374151'),
        bottom=Side(style='medium', color='374151')
    )
    ws.merge_cells('A4:J4')
    
    # Date de génération avec style discret
    ws['A5'] = f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A5'].font = Font(name='Arial', size=9, italic=True, color='9CA3AF')
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A5:J5')
    
    # Double ligne de séparation pour un effet professionnel
    for col in range(1, 11):
        # Première ligne
        cell = ws.cell(row=6, column=col)
        cell.fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        cell.border = Border(bottom=Side(style='thick', color='1F2937'))
        
        # Deuxième ligne plus fine
        cell2 = ws.cell(row=7, column=col)
        cell2.fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
    
    return 9  # Retourne la ligne de départ pour les données

def create_professional_data_section(ws, start_row, title, data_pairs):
    """Créer une section de données avec un style professionnel"""
    current_row = start_row
    
    # Titre de section avec style moderne
    ws[f'A{current_row}'] = title
    ws[f'A{current_row}'].style = "subheader_style"
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws.row_dimensions[current_row].height = 30
    
    current_row += 2
    
    # Données en deux colonnes pour un meilleur agencement
    for i in range(0, len(data_pairs), 2):
        # Première paire (colonne gauche)
        if i < len(data_pairs):
            label1, value1, style1 = data_pairs[i]
            ws[f'A{current_row}'] = label1
            ws[f'A{current_row}'].style = "label_style"
            ws[f'B{current_row}'] = value1
            apply_value_style(ws[f'B{current_row}'], style1)
        
        # Deuxième paire (colonne droite) si elle existe
        if i + 1 < len(data_pairs):
            label2, value2, style2 = data_pairs[i + 1]
            ws[f'D{current_row}'] = label2
            ws[f'D{current_row}'].style = "label_style"
            ws[f'E{current_row}'] = value2
            apply_value_style(ws[f'E{current_row}'], style2)
        
        current_row += 1
    
    return current_row + 1

def apply_value_style(cell, style_type):
    """Appliquer le style approprié selon le type"""
    if style_type == 'important':
        cell.style = "important_style"
    elif style_type == 'priority_high':
        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        cell.font = Font(bold=True, color='DC2626')
        cell.border = Border(
            left=Side(style='thin', color='F87171'),
            right=Side(style='thin', color='F87171'),
            top=Side(style='thin', color='F87171'),
            bottom=Side(style='thin', color='F87171')
        )
    elif style_type == 'priority_medium':
        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        cell.font = Font(bold=True, color='D97706')
        cell.border = Border(
            left=Side(style='thin', color='FBBF24'),
            right=Side(style='thin', color='FBBF24'),
            top=Side(style='thin', color='FBBF24'),
            bottom=Side(style='thin', color='FBBF24')
        )
    elif style_type == 'status_resolved':
        cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        cell.font = Font(bold=True, color='059669')
        cell.border = Border(
            left=Side(style='thin', color='34D399'),
            right=Side(style='thin', color='34D399'),
            top=Side(style='thin', color='34D399'),
            bottom=Side(style='thin', color='34D399')
        )
    else:
        cell.style = "important_style"

def create_professional_text_section(ws, start_row, title, content):
    """Créer une section de texte avec encadré professionnel"""
    current_row = start_row
    
    # Titre de section
    ws[f'A{current_row}'] = title
    ws[f'A{current_row}'].style = "subheader_style"
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws.row_dimensions[current_row].height = 25
    
    current_row += 1
    
    # Contenu avec encadré élégant
    ws[f'A{current_row}'] = content
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    ws[f'A{current_row}'].fill = PatternFill(start_color='FEFEFE', end_color='FEFEFE', fill_type='solid')
    ws[f'A{current_row}'].font = Font(name='Calibri', size=10, color='374151')
    ws[f'A{current_row}'].border = Border(
        left=Side(style='medium', color='D1D5DB'),
        right=Side(style='medium', color='D1D5DB'),
        top=Side(style='medium', color='D1D5DB'),
        bottom=Side(style='medium', color='D1D5DB')
    )
    ws.merge_cells(f'A{current_row}:F{current_row + 2}')
    ws.row_dimensions[current_row].height = 60
    
    return current_row + 4

def create_excel_export(pannes_data, single_panne=None):
    """Créer un fichier Excel avec design professionnel amélioré"""
    wb = openpyxl.Workbook()
    wb = create_excel_styles(wb)
    ws = wb.active
    
    if single_panne:
        ws.title = f"Panne_{single_panne['id']}"
        
        # En-tête professionnel avec logo centré
        start_row = add_excel_header(ws, f"RAPPORT DÉTAILLÉ - PANNE #{single_panne['id']}")
        
        # Section informations principales avec nouveau design
        current_row = start_row + 1
        info_data = [
            ('ID Panne:', f"#{single_panne['id']}", 'important'),
            ('Équipement:', single_panne['equipement'], 'normal'),
            ('Priorité:', single_panne['priorite'], get_priority_style(single_panne['priorite'])),
            ('État:', single_panne['etat'], get_status_style(single_panne['etat'])),
            ('Date de création:', single_panne['date_creation'], 'normal'),
            ('Créé par:', single_panne['username'], 'normal'),
        ]
        
        current_row = create_professional_data_section(ws, current_row, "INFORMATIONS GÉNÉRALES", info_data)
        
        # Sections de texte avec design amélioré
        current_row = create_professional_text_section(ws, current_row + 1, "DESCRIPTION DU PROBLÈME", single_panne['description'])
        current_row = create_professional_text_section(ws, current_row, "CAUSE IDENTIFIÉE", single_panne['cause'] or 'Non renseignée')
        current_row = create_professional_text_section(ws, current_row, "SOLUTION APPLIQUÉE", single_panne['solution'] or 'Non renseignée')
        current_row = create_professional_text_section(ws, current_row, "OBSERVATIONS", single_panne['observation'] or 'Aucune observation')
        
    else:
        ws.title = "Historique_Pannes"
        
        # En-tête professionnel
        start_row = add_excel_header(ws, "HISTORIQUE COMPLET DES PANNES")
        
        # Statistiques avec design moderne
        current_row = start_row + 1
        stats = calculate_stats(pannes_data)
        
        stats_data = [
            ('Total des pannes:', str(stats['total']), 'important'),
            ('En attente:', str(stats['en_attente']), 'normal'),
            ('En cours:', str(stats['en_cours']), 'normal'),
            ('Résolues:', str(stats['resolues']), 'status_resolved'),
            ('Taux de résolution:', f"{stats['taux_resolution']:.1f}%", 'important'),
            ('', '', 'normal')  # Pour équilibrer les colonnes
        ]
        
        current_row = create_professional_data_section(ws, current_row, "TABLEAU DE BORD", stats_data)
        
        # Tableau des pannes avec style professionnel
        current_row += 2
        ws[f'A{current_row}'] = "DÉTAIL DES PANNES"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws.row_dimensions[current_row].height = 30
        
        current_row += 2
        headers = ['ID', 'Équipement', 'Description', 'Priorité', 'État', 'Date', 'Créé par', 'Cause', 'Solution', 'Observations']
        
        # En-têtes du tableau avec style moderne
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.style = "header_style"
            ws.row_dimensions[current_row].height = 25
        
        # Données du tableau avec alternance de couleurs
        for row_idx, panne in enumerate(pannes_data, start=current_row + 1):
            # Couleur alternée pour les lignes
            is_even = (row_idx - current_row - 1) % 2 == 0
            base_color = 'FFFFFF' if is_even else 'F9FAFB'
            
            row_data = [
                f"#{panne['id']}",
                panne['equipement'],
                panne['description'][:50] + '...' if len(panne['description']) > 50 else panne['description'],
                panne['priorite'],
                panne['etat'],
                panne['date_creation'],
                panne['username'],
                panne['cause'] or 'Non renseignée',
                panne['solution'] or 'Non renseignée',
                panne['observation'] or 'Aucune'
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = PatternFill(start_color=base_color, end_color=base_color, fill_type="solid")
                cell.font = Font(name='Calibri', size=10, color='374151')
                cell.alignment = Alignment(vertical='center', horizontal='left')
                cell.border = Border(
                    left=Side(style='thin', color='E5E7EB'),
                    right=Side(style='thin', color='E5E7EB'),
                    top=Side(style='thin', color='E5E7EB'),
                    bottom=Side(style='thin', color='E5E7EB')
                )
            
            ws.row_dimensions[row_idx].height = 20
    
    # Ajuster les largeurs de colonnes de manière intelligente
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        # Largeurs optimisées selon le contenu
        if max_length < 10:
            adjusted_width = 12
        elif max_length < 20:
            adjusted_width = max_length + 3
        elif max_length < 40:
            adjusted_width = max_length + 2
        else:
            adjusted_width = 45
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def create_pdf_export(pannes_data, single_panne=None):
    """Créer un fichier PDF simple avec logo à gauche et maximum 2 couleurs"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Deux couleurs principales seulement
    PRIMARY_COLOR = colors.Color(0.12, 0.23, 0.54)  # Bleu foncé
    SECONDARY_COLOR = colors.black                    # Noir
    
    # Styles simples avec seulement 2 couleurs
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=0,  # Alignement à gauche
        textColor=PRIMARY_COLOR,
        fontName='Helvetica-Bold'
    )
    
    company_style = ParagraphStyle(
        'CompanyStyle',
        parent=styles['Normal'],
        fontSize=16,
        spaceAfter=5,
        alignment=0,  # Alignement à gauche
        textColor=PRIMARY_COLOR,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        alignment=0,  # Alignement à gauche
        textColor=SECONDARY_COLOR
    )
    
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceBefore=15,
        spaceAfter=8,
        textColor=PRIMARY_COLOR,
        fontName='Helvetica-Bold'
    )
    
    # Créer un tableau pour l'en-tête avec logo à gauche
    header_data = []
    
    # Logo à gauche dans une cellule
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        try:
            logo = Image(LOGO_PATH, width=0.8*inch, height=0.8*inch)
            # Créer une ligne avec logo à gauche et texte à droite
            header_table_data = [[logo, Paragraph(f"{COMPANY_NAME}<br/>{COMPANY_SUBTITLE}", company_style)]]
            header_table = Table(header_table_data, colWidths=[1*inch, 5*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            story.append(header_table)
        except Exception as e:
            print(f"Erreur logo PDF: {e}")
            # Si pas de logo, juste le texte à gauche
            story.append(Paragraph(COMPANY_NAME, company_style))
            story.append(Paragraph(COMPANY_SUBTITLE, subtitle_style))
    else:
        # Si pas de logo, juste le texte à gauche
        story.append(Paragraph(COMPANY_NAME, company_style))
        story.append(Paragraph(COMPANY_SUBTITLE, subtitle_style))
    
    # Ligne de séparation simple
    story.append(Spacer(1, 10))
    line_drawing = Drawing(500, 2)
    line_drawing.add(Rect(0, 0, 500, 2, fillColor=PRIMARY_COLOR, strokeColor=None))
    story.append(line_drawing)
    story.append(Spacer(1, 20))
    
    if single_panne:
        # Export détaillé d'une seule panne
        story.append(Paragraph(f"RAPPORT DÉTAILLÉ - PANNE #{single_panne['id']}", title_style))
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Informations principales - tableau simple
        story.append(Paragraph("INFORMATIONS GÉNÉRALES", section_style))
        
        main_data = [
            ['Équipement:', single_panne['equipement']],
            ['Priorité:', single_panne['priorite']],
            ['État:', single_panne['etat']],
            ['Date de création:', single_panne['date_creation']],
            ['Créé par:', single_panne['username']],
        ]
        
        main_table = Table(main_data, colWidths=[2*inch, 4*inch])
        main_table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, 0), (-1, -1), SECONDARY_COLOR),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, PRIMARY_COLOR),
        ]))
        
        story.append(main_table)
        story.append(Spacer(1, 20))
        
        # Description
        story.append(Paragraph("DESCRIPTION DU PROBLÈME", section_style))
        story.append(Paragraph(single_panne['description'], styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Cause
        story.append(Paragraph("CAUSE IDENTIFIÉE", section_style))
        story.append(Paragraph(single_panne['cause'] or 'Non renseignée', styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Solution
        story.append(Paragraph("SOLUTION APPLIQUÉE", section_style))
        story.append(Paragraph(single_panne['solution'] or 'Non renseignée', styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Observations
        story.append(Paragraph("OBSERVATIONS", section_style))
        story.append(Paragraph(single_panne['observation'] or 'Aucune observation', styles['Normal']))
        
    else:
        # Export de toutes les pannes
        story.append(Paragraph("HISTORIQUE COMPLET DES PANNES", title_style))
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Statistiques simples
        stats = calculate_stats(pannes_data)
        
        story.append(Paragraph("STATISTIQUES", section_style))
        
        stats_data = [
            ['Total des pannes', str(stats['total'])],
            ['En attente', str(stats['en_attente'])],
            ['En cours', str(stats['en_cours'])],
            ['Résolues', str(stats['resolues'])],
            ['Taux de résolution', f"{stats['taux_resolution']:.1f}%"],
        ]
        
        stats_table = Table(stats_data, colWidths=[2.5*inch, 1.5*inch])
        stats_table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, 0), (-1, -1), SECONDARY_COLOR),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, PRIMARY_COLOR),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 25))
        
        # Tableau des pannes - simple
        story.append(Paragraph("LISTE DES PANNES", section_style))
        
        table_data = [['ID', 'Équipement', 'Priorité', 'État', 'Date', 'Créé par']]
        
        for panne in pannes_data:
            table_data.append([
                f"#{panne['id']}",
                panne['equipement'][:20] + '...' if len(panne['equipement']) > 20 else panne['equipement'],
                panne['priorite'],
                panne['etat'],
                panne['date_creation'].split(' ')[0],
                panne['username'][:10] + '...' if len(panne['username']) > 10 else panne['username']
            ])
        
        table = Table(table_data, colWidths=[0.6*inch, 2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, 0), (-1, -1), SECONDARY_COLOR),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LINEBELOW', (0, 0), (-1, 0), 1, PRIMARY_COLOR),
            ('LINEBELOW', (0, 1), (-1, -1), 0.25, PRIMARY_COLOR),
        ]))
        
        story.append(table)
    
    # Pied de page simple
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,  # Centré
        textColor=SECONDARY_COLOR
    )
    story.append(Paragraph("ONEE-BO - Gestion des Pannes", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def get_performance_text(percentage):
    """Retourner une évaluation textuelle selon le pourcentage"""
    if percentage >= 95:
        return 'Performance excellente'
    elif percentage >= 85:
        return 'Performance très satisfaisante'
    elif percentage >= 75:
        return 'Performance satisfaisante'
    elif percentage >= 60:
        return 'Performance acceptable'
    elif percentage >= 40:
        return 'Performance à améliorer'
    else:
        return 'Performance critique'

# Fonctions utilitaires pour les styles
def get_priority_style(priority):
    """Retourner le style selon la priorité"""
    if priority == 'Haute':
        return 'priority_high'
    elif priority == 'Moyenne':
        return 'priority_medium'
    else:
        return 'normal'

def get_status_style(status):
    """Retourner le style selon le statut"""
    if status == 'Résolue':
        return 'status_resolved'
    else:
        return 'normal'

def get_status_color(status):
    """Retourner la couleur selon le statut pour Excel"""
    colors_map = {
        'Résolue': 'D1FAE5',      # Vert clair
        'En cours': 'DBEAFE',      # Bleu clair
        'En attente': 'FEF3C7',    # Jaune clair
        'default': 'FFFFFF'        # Blanc
    }
    return colors_map.get(status, colors_map['default'])

def get_performance_emoji(percentage):
    """Retourner un emoji selon le pourcentage de performance"""
    if percentage >= 90:
        return '🌟 Excellent'
    elif percentage >= 75:
        return '👍 Bon'
    elif percentage >= 50:
        return '⚠️ Moyen'
    else:
        return '🔴 Critique'

def calculate_stats(pannes_data):
    """Calculer les statistiques des pannes"""
    total = len(pannes_data)
    en_attente = len([p for p in pannes_data if p['etat'] == 'En attente'])
    en_cours = len([p for p in pannes_data if p['etat'] == 'En cours'])
    resolues = len([p for p in pannes_data if p['etat'] == 'Résolue'])
    
    taux_resolution = (resolues / total * 100) if total > 0 else 0
    
    return {
        'total': total,
        'en_attente': en_attente,
        'en_cours': en_cours,
        'resolues': resolues,
        'taux_resolution': taux_resolution
    }

def get_priority_style(priority):
    """Retourner le style selon la priorité"""
    if priority == 'Haute':
        return 'priority_high'
    elif priority == 'Moyenne':
        return 'priority_medium'
    else:
        return 'normal'

def get_status_style(status):
    """Retourner le style selon le statut"""
    if status == 'Résolue':
        return 'status_resolved'
    else:
        return 'normal'

def get_status_color(status):
    """Retourner la couleur selon le statut"""
    colors_map = {
        'Résolue': 'D1FAE5',      # Vert clair
        'En cours': 'DBEAFE',      # Bleu clair
        'En attente': 'FEF3C7',    # Jaune clair
        'default': 'F8FAFC'        # Gris très clair
    }
    return colors_map.get(status, colors_map['default'])

def get_performance_emoji(percentage):
    """Retourner un emoji selon le pourcentage de performance"""
    if percentage >= 90:
        return '🌟'
    elif percentage >= 75:
        return '👍'
    elif percentage >= 50:
        return '⚠️'
    else:
        return '🔴'

def create_excel_export(pannes_data, single_panne=None):
    """Créer un fichier Excel avec design amélioré pour le dessalement"""
    wb = openpyxl.Workbook()
    wb = create_excel_styles(wb)
    ws = wb.active
    
    if single_panne:
        ws.title = f"Panne_{single_panne['id']}"
        
        # En-tête décoratif
        start_row = add_excel_header(ws, f"RAPPORT DÉTAILLÉ - PANNE #{single_panne['id']}")
        
        # Section informations principales
        current_row = start_row + 2
        ws[f'A{current_row}'] = "INFORMATIONS GÉNÉRALES"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 2
        details = [
            ('ID Panne:', f"#{single_panne['id']}", 'important'),
            ('Équipement:', single_panne['equipement'], 'normal'),
            ('Priorité:', single_panne['priorite'], get_priority_style(single_panne['priorite'])),
            ('État:', single_panne['etat'], get_status_style(single_panne['etat'])),
            ('Date de création:', single_panne['date_creation'], 'normal'),
            ('Créé par:', single_panne['username'], 'normal'),
        ]
        
        for label, value, style_type in details:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True, color=THEME_COLORS['header'])
            ws[f'B{current_row}'] = value
            
            if style_type == 'important':
                ws[f'B{current_row}'].style = "important_style"
            elif style_type == 'priority_high':
                ws[f'B{current_row}'].fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                ws[f'B{current_row}'].font = Font(bold=True, color='DC2626')
            elif style_type == 'priority_medium':
                ws[f'B{current_row}'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                ws[f'B{current_row}'].font = Font(bold=True, color='D97706')
            elif style_type == 'status_resolved':
                ws[f'B{current_row}'].fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                ws[f'B{current_row}'].font = Font(bold=True, color='059669')
            
            current_row += 1
        
        # Section description
        current_row += 2
        ws[f'A{current_row}'] = "DESCRIPTION DU PROBLÈME"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = single_panne['description']
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row}:D{current_row + 2}')
        ws.row_dimensions[current_row].height = 60
        
        # Section cause
        current_row += 4
        ws[f'A{current_row}'] = "CAUSE IDENTIFIÉE"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = single_panne['cause'] or 'Non renseignée'
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row}:D{current_row + 1}')
        
        # Section solution
        current_row += 3
        ws[f'A{current_row}'] = "SOLUTION APPLIQUÉE"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = single_panne['solution'] or 'Non renseignée'
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row}:D{current_row + 1}')
        
        # Section observations
        current_row += 3
        ws[f'A{current_row}'] = "OBSERVATIONS"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = single_panne['observation'] or 'Aucune observation'
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{current_row}:D{current_row + 1}')
        
    else:
        ws.title = "Historique_Pannes"
        
        # En-tête décoratif
        start_row = add_excel_header(ws, "HISTORIQUE COMPLET DES PANNES")
        
        # Statistiques
        current_row = start_row + 1
        stats = calculate_stats(pannes_data)
        
        ws[f'A{current_row}'] = "STATISTIQUES GÉNÉRALES"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 2
        stats_data = [
            ('Total des pannes:', stats['total']),
            ('En attente:', stats['en_attente']),
            ('En cours:', stats['en_cours']),
            ('Résolues:', stats['resolues']),
            ('Taux de résolution:', f"{stats['taux_resolution']:.1f}%")
        ]
        
        for label, value in stats_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = "important_style"
            current_row += 1
        
        # Tableau des pannes
        current_row += 2
        ws[f'A{current_row}'] = "DÉTAIL DES PANNES"
        ws[f'A{current_row}'].style = "subheader_style"
        ws.merge_cells(f'A{current_row}:J{current_row}')
        
        current_row += 2
        headers = ['ID', 'Équipement', 'Description', 'Priorité', 'État', 'Date', 'Créé par', 'Cause', 'Solution', 'Observations']
        
        # En-têtes du tableau
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.style = "header_style"
        
        # Données du tableau
        for row_idx, panne in enumerate(pannes_data, start=current_row + 1):
            ws.cell(row=row_idx, column=1, value=f"#{panne['id']}")
            ws.cell(row=row_idx, column=2, value=panne['equipement'])
            ws.cell(row=row_idx, column=3, value=panne['description'][:50] + '...' if len(panne['description']) > 50 else panne['description'])
            ws.cell(row=row_idx, column=4, value=panne['priorite'])
            ws.cell(row=row_idx, column=5, value=panne['etat'])
            ws.cell(row=row_idx, column=6, value=panne['date_creation'])
            ws.cell(row=row_idx, column=7, value=panne['username'])
            ws.cell(row=row_idx, column=8, value=panne['cause'] or 'Non renseignée')
            ws.cell(row=row_idx, column=9, value=panne['solution'] or 'Non renseignée')
            ws.cell(row=row_idx, column=10, value=panne['observation'] or 'Aucune')
            
            # Couleur selon l'état
            fill_color = get_status_color(panne['etat'])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                ws.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
    
    # Ajuster les largeurs de colonnes
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        adjusted_width = min(max(max_length + 2, 15), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

# Routes Flask

@app.route('/')
def index():
    if 'user_id' in session:
        if session['role'] == 'admin':
            return redirect(url_for('dashboard'))
        else:
            return redirect(url_for('historique'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        user = conn.execute(
            'SELECT * FROM users WHERE username = ? AND password = ?',
            (username, password)
        ).fetchone()
        conn.close()
        
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash(f'Connexion réussie ! Bienvenue {user["username"]}', 'success')
            
            if user['role'] == 'admin':
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('historique'))
        else:
            flash('Nom d\'utilisateur ou mot de passe incorrect', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Déconnexion réussie', 'success')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
@admin_required
def dashboard():
    conn = get_db_connection()
    
    total_pannes = conn.execute('SELECT COUNT(*) as count FROM pannes').fetchone()['count']
    pannes_en_attente = conn.execute('SELECT COUNT(*) as count FROM pannes WHERE etat = "En attente"').fetchone()['count']
    pannes_resolues = conn.execute('SELECT COUNT(*) as count FROM pannes WHERE etat = "Résolue"').fetchone()['count']
    pannes_en_cours = conn.execute('SELECT COUNT(*) as count FROM pannes WHERE etat = "En cours"').fetchone()['count']
    
    pannes_recentes = conn.execute(
        'SELECT * FROM pannes ORDER BY date_creation DESC LIMIT 5'
    ).fetchall()
    
    conn.close()
    
    stats = {
        'total': total_pannes,
        'en_attente': pannes_en_attente,
        'resolues': pannes_resolues,
        'en_cours': pannes_en_cours
    }
    
    return render_template('dashboard.html', stats=stats, pannes_recentes=pannes_recentes)

@app.route('/ajouter_panne', methods=['GET', 'POST'])
@login_required
def ajouter_panne():
    if request.method == 'POST':
        equipement = request.form['equipement']
        description = request.form['description']
        priorite = request.form['priorite']
        etat = request.form['etat']
        cause = request.form['cause']
        solution = request.form['solution']
        observation = request.form['observation']
        date_creation = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        conn = get_db_connection()
        conn.execute(
            '''INSERT INTO pannes (equipement, description, priorite, etat, date_creation, 
               cause, solution, observation, user_id) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (equipement, description, priorite, etat, date_creation, cause, solution, 
             observation, session['user_id'])
        )
        conn.commit()
        conn.close()
        
        flash('Panne ajoutée avec succès !', 'success')
        return redirect(url_for('historique'))
    
    return render_template('ajouter_panne.html')

@app.route('/historique')
@login_required
def historique():
    conn = get_db_connection()
    pannes = conn.execute(
        '''SELECT p.*, u.username FROM pannes p 
           JOIN users u ON p.user_id = u.id 
           ORDER BY p.date_creation DESC'''
    ).fetchall()
    conn.close()
    
    return render_template('historique.html', pannes=pannes)

@app.route('/modifier_panne/<int:panne_id>', methods=['GET', 'POST'])
@login_required
def modifier_panne(panne_id):
    conn = get_db_connection()
    
    if request.method == 'POST':
        equipement = request.form['equipement']
        description = request.form['description']
        priorite = request.form['priorite']
        etat = request.form['etat']
        cause = request.form['cause']
        solution = request.form['solution']
        observation = request.form['observation']
        
        conn.execute(
            '''UPDATE pannes SET equipement=?, description=?, priorite=?, etat=?, 
               cause=?, solution=?, observation=? WHERE id=?''',
            (equipement, description, priorite, etat, cause, solution, observation, panne_id)
        )
        conn.commit()
        conn.close()
        
        flash('Panne modifiée avec succès !', 'success')
        return redirect(url_for('historique'))
    
    panne = conn.execute('SELECT * FROM pannes WHERE id = ?', (panne_id,)).fetchone()
    conn.close()
    
    if not panne:
        flash('Panne non trouvée', 'error')
        return redirect(url_for('historique'))
    
    return render_template('ajouter_panne.html', panne=panne, mode='modifier')

@app.route('/export_excel')
@login_required
@admin_required
def export_excel():
    """Exporter toutes les pannes en Excel avec design amélioré"""
    conn = get_db_connection()
    pannes = conn.execute(
        '''SELECT p.*, u.username FROM pannes p 
           JOIN users u ON p.user_id = u.id 
           ORDER BY p.date_creation DESC'''
    ).fetchall()
    conn.close()
    
    wb = create_excel_export(pannes)
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Gestion_Pannes_ONEE-BO_Historique_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    flash('Export Excel généré avec succès !', 'success')
    return response

@app.route('/export_pdf')
@login_required
@admin_required
def export_pdf():
    """Exporter toutes les pannes en PDF"""
    conn = get_db_connection()
    pannes = conn.execute(
        '''SELECT p.*, u.username FROM pannes p 
           JOIN users u ON p.user_id = u.id 
           ORDER BY p.date_creation DESC'''
    ).fetchall()
    conn.close()
    
    pdf_buffer = create_pdf_export(pannes)
    
    response = make_response(pdf_buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=historique_pannes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    flash('Export PDF généré avec succès !', 'success')
    return response

@app.route('/export_panne_excel/<int:panne_id>')
@login_required
@admin_required
def export_panne_excel(panne_id):
    """Exporter une panne spécifique en Excel avec design amélioré"""
    conn = get_db_connection()
    panne = conn.execute(
        '''SELECT p.*, u.username FROM pannes p 
           JOIN users u ON p.user_id = u.id 
           WHERE p.id = ?''', (panne_id,)
    ).fetchone()
    conn.close()
    
    if not panne:
        flash('Panne non trouvée', 'error')
        return redirect(url_for('historique'))
    
    wb = create_excel_export(None, panne)
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Gestion_Pannes_ONEE-BO_{panne_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    flash(f'Export Excel de la panne #{panne_id} généré !', 'success')
    return response

@app.route('/export_panne_pdf/<int:panne_id>')
@login_required
@admin_required
def export_panne_pdf(panne_id):
    """Exporter une panne spécifique en PDF"""
    conn = get_db_connection()
    panne = conn.execute(
        '''SELECT p.*, u.username FROM pannes p 
           JOIN users u ON p.user_id = u.id 
           WHERE p.id = ?''', (panne_id,)
    ).fetchone()
    conn.close()
    
    if not panne:
        flash('Panne non trouvée', 'error')
        return redirect(url_for('historique'))
    
    pdf_buffer = create_pdf_export(None, panne)
    
    response = make_response(pdf_buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=panne_{panne_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    flash(f'Export PDF de la panne #{panne_id} généré !', 'success')
    return response


if __name__ == '__main__':
    app.run(debug=True)